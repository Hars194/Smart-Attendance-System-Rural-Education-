import os
import io
import re
import csv
from urllib.parse import quote_plus
from datetime import datetime, date, timedelta, time
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, session
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask_bcrypt import Bcrypt
from flask_cors import CORS
from werkzeug.utils import secure_filename
from sqlalchemy import func
from sqlalchemy.exc import IntegrityError
from fpdf import FPDF
from openpyxl import Workbook

# --- IMPORT TRANSLATIONS ---
from languages import TRANSLATIONS

# --- CONFIGURATION ---
app = Flask(__name__)
app.config['SECRET_KEY'] = 'secure-key-change-in-production'

# ---------------------------------------------------------
# DATABASE CONFIGURATION (SQLITE)
# ---------------------------------------------------------

basedir = os.path.abspath(os.path.dirname(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(basedir, 'school_system.db')

app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_recycle': 280,
    'pool_pre_ping': True
}

app.config['UPLOAD_FOLDER'] = 'static/uploads'
app.config['ALLOWED_EXTENSIONS'] = {'png', 'jpg', 'jpeg'}

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# --- GLOBAL VARIABLES ---
DATA_LAST_UPDATED = datetime.now().timestamp() # For Auto-Sync
RATE_PRIMARY = 6.78       # Cost for Class 1-5
RATE_UPPER_PRIMARY = 10.17 # Cost for Class 6-8

db = SQLAlchemy(app)
bcrypt = Bcrypt(app)
CORS(app) 

login_manager = LoginManager(app)
login_manager.login_view = 'login_page'

# --- LANGUAGE SETTINGS ---
@app.context_processor
def inject_language():
    lang_code = session.get('lang', 'en')
    return dict(t=TRANSLATIONS.get(lang_code, TRANSLATIONS['en']), current_lang=lang_code)

@app.route('/set_language/<lang_code>')
def set_language(lang_code):
    if lang_code in TRANSLATIONS:
        session['lang'] = lang_code
    return redirect(request.referrer or url_for('login_page'))

# --- DATABASE MODELS ---

class School(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(150), unique=True, nullable=False)
    # --- NEW SETTINGS FOR ATTENDANCE ---
    start_time = db.Column(db.String(10), default="08:00") # "HH:MM" 24-hour format
    duration = db.Column(db.Integer, default=60)           # Duration in minutes
    # -----------------------------------
    users = db.relationship('User', backref='school', lazy=True, cascade="all, delete-orphan")
    students = db.relationship('Student', backref='school', lazy=True, cascade="all, delete-orphan")
    logs = db.relationship('AttendanceLog', backref='school', lazy=True, cascade="all, delete-orphan")

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(150), unique=True, nullable=False)
    password = db.Column(db.String(150), nullable=False)
    role = db.Column(db.String(20), nullable=False) 
    school_id = db.Column(db.Integer, db.ForeignKey('school.id'), nullable=True)
    assigned_class = db.Column(db.String(50), nullable=True) 

class Student(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    rfid_uid = db.Column(db.String(50), unique=True, nullable=False)
    image_file = db.Column(db.String(100), nullable=False)
    classroom = db.Column(db.String(50), nullable=False) 
    school_id = db.Column(db.Integer, db.ForeignKey('school.id'), nullable=False)

class AttendanceLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_name = db.Column(db.String(100))
    student_uid = db.Column(db.String(50))
    status = db.Column(db.String(20)) 
    place = db.Column(db.String(100))
    timestamp = db.Column(db.DateTime, default=datetime.now)
    school_id = db.Column(db.Integer, db.ForeignKey('school.id'), nullable=True)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# --- HELPER FUNCTIONS ---

def get_filtered_query(base_query, start_date, end_date, filter_id=None, filter_type=None):
    if start_date and end_date:
        start = datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
        base_query = base_query.filter(AttendanceLog.timestamp >= start, AttendanceLog.timestamp < end)
    elif start_date:
        start = datetime.strptime(start_date, '%Y-%m-%d')
        end = start + timedelta(days=1)
        base_query = base_query.filter(AttendanceLog.timestamp >= start, AttendanceLog.timestamp < end)
        
    if filter_type == 'school' and filter_id:
        base_query = base_query.filter(AttendanceLog.school_id == int(filter_id))
    elif filter_type == 'class' and filter_id:
        base_query = base_query.join(Student, AttendanceLog.student_uid == Student.rfid_uid)\
                               .filter(Student.classroom == filter_id)
    return base_query.order_by(AttendanceLog.timestamp.desc())

def get_mdm_category(classroom_str):
    match = re.search(r'\d+', str(classroom_str))
    if match:
        class_num = int(match.group())
        if 1 <= class_num <= 5: return 'primary'
        elif 6 <= class_num <= 8: return 'upper'
    return 'none'

def trigger_auto_absent(school_id):
    """
    Checks all students in the school. If they haven't marked attendance today,
    mark them as 'ABSENT' immediately.
    """
    try:
        today_str = date.today().strftime('%Y-%m-%d')
        today_dt = datetime.now()
        
        # 1. Get all students for this school
        all_students = Student.query.filter_by(school_id=school_id).all()
        
        # 2. Get students who are already present/marked today
        present_logs = AttendanceLog.query.filter(
            AttendanceLog.school_id == school_id,
            func.date(AttendanceLog.timestamp) == today_str
        ).all()
        
        present_uids = {log.student_uid for log in present_logs}
        
        # 3. Mark missing students as absent
        count = 0
        for student in all_students:
            if student.rfid_uid not in present_uids:
                absent_log = AttendanceLog(
                    student_name=student.name,
                    student_uid=student.rfid_uid,
                    status='ABSENT',
                    place='Auto-System',
                    timestamp=today_dt,
                    school_id=school_id
                )
                db.session.add(absent_log)
                count += 1
        
        if count > 0:
            db.session.commit()
            print(f"DEBUG: Marked {count} students Absent for School ID {school_id}")
            
            # Trigger update signal for dashboards
            global DATA_LAST_UPDATED
            DATA_LAST_UPDATED = datetime.now().timestamp()
            
    except Exception as e:
        print(f"Error in auto-absent: {e}")

# --- AUTH ROUTES ---
@app.route('/')
def login_page():
    if current_user.is_authenticated: return handle_redirect(current_user)
    return render_template('login.html')

@app.route('/api/auth/login', methods=['POST'])
def api_login():
    data = request.json
    user = User.query.filter_by(username=data.get('username')).first()
    if user and bcrypt.check_password_hash(user.password, data.get('password')):
        login_user(user)
        redirect_url = ""
        if user.role == 'super_admin': redirect_url = "/super-admin/dashboard"
        elif user.role == 'principal': redirect_url = f"/school/{user.school_id}/dashboard"
        elif user.role == 'teacher': redirect_url = f"/school/{user.school_id}/teacher-dashboard"
        return jsonify({"status": "success", "redirect_url": redirect_url}), 200
    return jsonify({"status": "error", "message": "Invalid Credentials"}), 401

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login_page'))

def handle_redirect(user):
    if user.role == 'super_admin': return redirect('/super-admin/dashboard')
    if user.role == 'principal': return redirect(f'/school/{user.school_id}/dashboard')
    if user.role == 'teacher': return redirect(f'/school/{user.school_id}/teacher-dashboard')
    return redirect('/')

# --- MANAGEMENT ROUTES ---
@app.route('/create_user', methods=['POST'])
@login_required
def create_user():
    username = request.form.get('username')
    password = request.form.get('password')
    hashed_pw = bcrypt.generate_password_hash(password).decode('utf-8')

    if current_user.role == 'principal':
        assigned_class = request.form.get('assigned_class')
        db.session.add(User(username=username, password=hashed_pw, role='teacher', school_id=current_user.school_id, assigned_class=assigned_class))
        flash(f'Teacher created for Class {assigned_class}!', 'success')
    elif current_user.role == 'super_admin':
        school_id = request.form.get('school_id')
        db.session.add(User(username=username, password=hashed_pw, role='principal', school_id=school_id))
        flash('Principal Account Created!', 'success')
    
    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        flash("Error: Username already exists!", "danger")
        
    return redirect(request.referrer)

@app.route('/edit_user/<int:user_id>', methods=['POST'])
@login_required
def edit_user(user_id):
    user = User.query.get_or_404(user_id)
    allowed = False
    if current_user.role == 'super_admin' and user.role == 'principal': allowed = True
    if current_user.role == 'principal' and user.role == 'teacher' and user.school_id == current_user.school_id: allowed = True
    
    if allowed:
        user.username = request.form.get('username')
        if request.form.get('password'): 
            user.password = bcrypt.generate_password_hash(request.form.get('password')).decode('utf-8')
        if request.form.get('assigned_class'):
            user.assigned_class = request.form.get('assigned_class')
        db.session.commit()
        flash("User updated", "success")
    else:
        flash("Unauthorized", "danger")
    return redirect(request.referrer)

@app.route('/delete_user/<int:user_id>')
@login_required
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    if (current_user.role == 'super_admin' and user.role == 'principal') or \
       (current_user.role == 'principal' and user.role == 'teacher' and user.school_id == current_user.school_id):
        db.session.delete(user)
        db.session.commit()
        flash("User deleted", "warning")
    else:
        flash("Unauthorized", "danger")
    return redirect(request.referrer)

@app.route('/create_school', methods=['POST'])
@login_required
def create_school():
    if current_user.role == 'super_admin':
        db.session.add(School(name=request.form.get('name')))
        db.session.commit()
    return redirect('/super-admin/dashboard')

@app.route('/delete_school/<int:school_id>')
@login_required
def delete_school(school_id):
    if current_user.role != 'super_admin': return "Unauthorized", 403
    school = School.query.get_or_404(school_id)
    db.session.delete(school)
    db.session.commit()
    flash(f"School '{school.name}' deleted.", "danger")
    return redirect('/super-admin/dashboard')

# --- MODIFIED: UPDATE TIME SETTINGS (EXTENDED TO 24 HR) ---
@app.route('/update_time_settings', methods=['POST'])
@login_required
def update_time_settings():
    if current_user.role != 'principal':
        return redirect('/')
    
    school = School.query.get(current_user.school_id)
    if school:
        school.start_time = request.form.get('start_time')
        
        # Validation: Extended to 1440 minutes (24 Hours) for Testing
        try:
            duration_val = int(request.form.get('duration'))
            if duration_val > 1440:
                flash("Error: Duration cannot exceed 1440 minutes (24 Hours).", "danger")
            elif duration_val < 1:
                flash("Error: Duration must be at least 1 minute.", "danger")
            else:
                school.duration = duration_val
                db.session.commit()
                flash("Attendance time settings updated!", "success")
        except ValueError:
            flash("Invalid duration format.", "danger")
            
    return redirect(f'/school/{current_user.school_id}/dashboard')


# --- DASHBOARDS ---
@app.route('/super-admin/dashboard')
@login_required
def super_admin_dashboard():
    if current_user.role != 'super_admin': return redirect('/')
    schools = School.query.all()
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    selected_school_id = request.args.get('school_id')
    
    query = AttendanceLog.query
    logs = get_filtered_query(query, start_date, end_date, selected_school_id, 'school').limit(100).all()

    today_str = str(date.today())
    stats = []
    for sch in schools:
        total = Student.query.filter_by(school_id=sch.id).count()
        present = db.session.query(AttendanceLog.student_uid).filter(
            AttendanceLog.school_id == sch.id, AttendanceLog.status == 'PRESENT',
            func.date(AttendanceLog.timestamp) == today_str
        ).distinct().count()
        percentage = round((present / total) * 100, 1) if total > 0 else 0
        principal = User.query.filter_by(school_id=sch.id, role='principal').first()
        p_name = principal.username if principal else None
        p_id = principal.id if principal else None
        stats.append({'id': sch.id, 'name': sch.name, 'percentage': percentage, 'principal': p_name, 'p_id': p_id})

    return render_template('dashboard_super.html', schools=schools, logs=logs, stats=stats, start_date=start_date, end_date=end_date, selected_school_id=selected_school_id)

@app.route('/school/<int:school_id>/dashboard')
@login_required
def principal_dashboard(school_id):
    if current_user.role != 'principal' or current_user.school_id != school_id: return "Unauthorized", 403
    school = School.query.get_or_404(school_id)
    students = Student.query.filter_by(school_id=school_id).all()
    teachers = User.query.filter_by(school_id=school_id, role='teacher').all()
    existing_classes = db.session.query(Student.classroom).filter_by(school_id=school_id).distinct().all()
    class_list = [c[0] for c in existing_classes]
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    selected_class = request.args.get('classroom')
    base_query = AttendanceLog.query.filter_by(school_id=school_id)
    logs = get_filtered_query(base_query, start_date, end_date, selected_class, 'class').all()
    today_str = str(date.today())
    class_stats = []
    for cls_name in class_list:
        total = Student.query.filter_by(school_id=school_id, classroom=cls_name).count()
        present = db.session.query(AttendanceLog.student_uid).join(Student, AttendanceLog.student_uid == Student.rfid_uid).filter(
            AttendanceLog.school_id == school_id, Student.classroom == cls_name,
            AttendanceLog.status == 'PRESENT', func.date(AttendanceLog.timestamp) == today_str
        ).distinct().count()
        percentage = round((present / total) * 100, 1) if total > 0 else 0
        class_stats.append({'name': cls_name, 'percentage': percentage})
    return render_template('dashboard_principal.html', school=school, students=students, teachers=teachers, logs=logs, class_stats=class_stats, class_list=class_list, start_date=start_date, end_date=end_date, selected_class=selected_class)

@app.route('/school/<int:school_id>/teacher-dashboard')
@login_required
def teacher_dashboard(school_id):
    if current_user.role != 'teacher' or current_user.school_id != school_id: return "Unauthorized", 403
    school = School.query.get_or_404(school_id)
    my_class = current_user.assigned_class
    students = Student.query.filter_by(school_id=school_id, classroom=my_class).all()
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    base_query = db.session.query(AttendanceLog).join(Student, AttendanceLog.student_uid == Student.rfid_uid)\
        .filter(AttendanceLog.school_id == school_id, Student.classroom == my_class)\
        .order_by(AttendanceLog.timestamp.desc())
    logs = get_filtered_query(base_query, start_date, end_date).all()
    return render_template('dashboard_teacher.html', school=school, logs=logs, students=students, my_class=my_class, start_date=start_date, end_date=end_date)

# --- MID-DAY MEAL (MDM) ROUTES ---
@app.route('/super-admin/mdm-report')
@login_required
def mdm_report_page():
    if current_user.role != 'super_admin': return redirect('/')

    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not start_date or not end_date:
        today = date.today()
        start_date = today.replace(day=1).strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')

    s_dt = datetime.strptime(start_date, '%Y-%m-%d')
    e_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)

    logs = AttendanceLog.query.filter(
        AttendanceLog.status == 'PRESENT',
        AttendanceLog.timestamp >= s_dt,
        AttendanceLog.timestamp < e_dt
    ).all()

    mdm_data = {
        'total_cost': 0.0,
        'primary_count': 0,
        'upper_count': 0,
        'school_breakdown': {},
        'daily_trend': {} 
    }

    student_cache = {s.rfid_uid: s.classroom for s in Student.query.all()}

    for log in logs:
        classroom = student_cache.get(log.student_uid)
        if not classroom: continue

        category = get_mdm_category(classroom)
        cost = 0
        
        if category == 'primary':
            cost = RATE_PRIMARY
            mdm_data['primary_count'] += 1
        elif category == 'upper':
            cost = RATE_UPPER_PRIMARY
            mdm_data['upper_count'] += 1
        
        if cost > 0:
            mdm_data['total_cost'] += cost
            
            sch_id = log.school_id
            if sch_id not in mdm_data['school_breakdown']:
                school_name = School.query.get(sch_id).name if sch_id else "Unknown"
                mdm_data['school_breakdown'][sch_id] = {'name': school_name, 'cost': 0, 'p_count': 0, 'u_count': 0}
            
            mdm_data['school_breakdown'][sch_id]['cost'] += cost
            if category == 'primary': mdm_data['school_breakdown'][sch_id]['p_count'] += 1
            else: mdm_data['school_breakdown'][sch_id]['u_count'] += 1

            log_date = log.timestamp.strftime('%Y-%m-%d')
            if log_date not in mdm_data['daily_trend']:
                mdm_data['daily_trend'][log_date] = 0
            mdm_data['daily_trend'][log_date] += cost

    graph_labels = sorted(mdm_data['daily_trend'].keys())
    graph_values = [round(mdm_data['daily_trend'][d], 2) for d in graph_labels]

    return render_template('mdm_report.html', mdm_data=mdm_data, graph_labels=graph_labels, graph_values=graph_values, start_date=start_date, end_date=end_date)

@app.route('/export_mdm_report')
@login_required
def export_mdm_report():
    if current_user.role != 'super_admin': return "Unauthorized", 403
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    if not start_date or not end_date:
        today = date.today()
        start_date = today.replace(day=1).strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')

    s_dt = datetime.strptime(start_date, '%Y-%m-%d')
    e_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)

    logs = AttendanceLog.query.filter(AttendanceLog.status == 'PRESENT', AttendanceLog.timestamp >= s_dt, AttendanceLog.timestamp < e_dt).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "MDM Cost Report"
    ws.append(["Date", "School", "Student UID", "Category", "Rate", "Cost"])
    
    total = 0
    for log in logs:
        student = Student.query.filter_by(rfid_uid=log.student_uid).first()
        if student:
            cat = get_mdm_category(student.classroom)
            cost = 0
            if cat == 'primary': cost = RATE_PRIMARY
            elif cat == 'upper': cost = RATE_UPPER_PRIMARY
            
            if cost > 0:
                school_name = log.school.name if log.school else "Unknown"
                ws.append([log.timestamp.strftime('%Y-%m-%d'), school_name, log.student_uid, cat, cost, cost])
                total += cost

    ws.append(["", "", "", "", "TOTAL", total])
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="MDM_Cost_Report.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# --- STUDENT & LOG MANAGEMENT ---
@app.route('/add_student', methods=['GET', 'POST'])
@login_required
def add_student():
    if current_user.role not in ['principal', 'super_admin', 'teacher']: return "Unauthorized", 403
    
    if request.method == 'POST':
        if 'image' in request.files:
            file = request.files['image']
            filename = secure_filename(f"{request.form.get('rfid')}_{request.form.get('name')}.jpg")
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
        else:
            filename = "default.jpg"

        classroom = current_user.assigned_class if current_user.role == 'teacher' else request.form.get('classroom')
        new_student = Student(
            name=request.form.get('name'), 
            rfid_uid=request.form.get('rfid'), 
            classroom=classroom, 
            image_file=filename, 
            school_id=current_user.school_id
        )

        try:
            db.session.add(new_student)
            db.session.commit()
            
            global DATA_LAST_UPDATED
            DATA_LAST_UPDATED = datetime.now().timestamp()

            flash("Student added successfully!", "success")
            if current_user.role == 'teacher':
                return redirect(f'/school/{current_user.school_id}/teacher-dashboard')
            else:
                return redirect(f'/school/{current_user.school_id}/dashboard')
                
        except IntegrityError:
            db.session.rollback()
            flash(f"Error: A student with RFID '{request.form.get('rfid')}' already exists!", "danger")
            return redirect(url_for('add_student'))

    prefill = current_user.assigned_class if current_user.role == 'teacher' else None
    return render_template('add_student.html', student=None, prefill_class=prefill)

@app.route('/edit_student/<int:student_id>', methods=['GET', 'POST'])
@login_required
def edit_student(student_id):
    student = Student.query.get_or_404(student_id)
    if request.method == 'POST':
        student.name = request.form.get('name')
        student.rfid_uid = request.form.get('rfid')
        if current_user.role != 'teacher': student.classroom = request.form.get('classroom')
        if request.files['image']: 
            file = request.files['image']
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(f"{student.rfid_uid}_{student.name}.jpg")))
        
        try:
            db.session.commit()
            
            global DATA_LAST_UPDATED
            DATA_LAST_UPDATED = datetime.now().timestamp()
            
            flash("Student updated!", "success")
        except IntegrityError:
            db.session.rollback()
            flash("Error: That RFID or Data conflicts with another student.", "danger")

        return redirect(request.referrer)
    return render_template('add_student.html', student=student)

@app.route('/delete_student/<int:student_id>')
@login_required
def delete_student(student_id):
    student = Student.query.get_or_404(student_id)
    try: os.remove(os.path.join(app.config['UPLOAD_FOLDER'], student.image_file))
    except: pass
    db.session.delete(student)
    db.session.commit()
    
    global DATA_LAST_UPDATED
    DATA_LAST_UPDATED = datetime.now().timestamp()
    
    return redirect(request.referrer)

@app.route('/delete_log/<int:log_id>')
@login_required
def delete_log(log_id):
    db.session.delete(AttendanceLog.query.get_or_404(log_id)); db.session.commit()
    return redirect(request.referrer)

@app.route('/export_report')
@login_required
def export_report():
    fmt = request.args.get('format')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    query = AttendanceLog.query
    if current_user.role == 'principal': query = query.filter_by(school_id=current_user.school_id)
    elif current_user.role == 'teacher':
        my_class = current_user.assigned_class
        query = db.session.query(AttendanceLog).join(Student, AttendanceLog.student_uid == Student.rfid_uid)\
            .filter(AttendanceLog.school_id == current_user.school_id, Student.classroom == my_class)
    if start_date: query = query.filter(AttendanceLog.timestamp >= datetime.strptime(start_date, '%Y-%m-%d'))
    if end_date: query = query.filter(AttendanceLog.timestamp < datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1))
    logs = query.order_by(AttendanceLog.timestamp.desc()).all()

    if fmt == 'excel':
        wb = Workbook(); ws = wb.active; ws.append(["Date","Time","Name","UID","Status"])
        for l in logs: ws.append([l.timestamp.strftime('%Y-%m-%d'), l.timestamp.strftime('%H:%M'), l.student_name, l.student_uid, l.status])
        buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name="Report.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    elif fmt == 'pdf':
        pdf = FPDF(); pdf.add_page(); pdf.set_font("Arial", size=12)
        pdf.cell(200, 10, txt=f"Report {date.today()}", ln=True, align='C'); pdf.ln(10)
        pdf.cell(40, 10, "Date", 1); pdf.cell(50, 10, "Name", 1); pdf.cell(40, 10, "Status", 1); pdf.ln()
        for log in logs:
            pdf.cell(40, 10, log.timestamp.strftime('%Y-%m-%d'), 1)
            pdf.cell(50, 10, log.student_name, 1)
            pdf.cell(40, 10, log.status, 1); pdf.ln()
        buffer = io.BytesIO(); pdf_bytes = pdf.output(dest='S').encode('latin-1'); buffer.write(pdf_bytes); buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name="Report.pdf", mimetype='application/pdf')
    return "Invalid"

# --- API ENDPOINTS (FOR RASPBERRY PI) ---
@app.route('/api/sync/status', methods=['GET'])
def get_sync_status():
    global DATA_LAST_UPDATED
    return jsonify({"last_updated": DATA_LAST_UPDATED})

@app.route('/api/students', methods=['GET'])
def get_all_students():
    return jsonify([{"name": s.name, "uid": s.rfid_uid, "image_file": s.image_file} for s in Student.query.all()])

# --- UPDATED ATTENDANCE API WITH TIME CHECK ---
@app.route('/api/update', methods=['POST'])
def receive_attendance():
    try:
        data = request.json
        uid = data.get('uid')
        input_status = data.get('status') # 'PRESENT' or 'MISMATCH' sent by Pi
        
        student = Student.query.filter_by(rfid_uid=uid).first()
        
        if not student:
            return jsonify({"status": "error", "message": "Student not found"}), 404
            
        school = School.query.get(student.school_id)
        
        # --- TIME VALIDATION LOGIC ---
        now = datetime.now()
        
        # Parse School Start Time (Stored as String "HH:MM")
        try:
            sch_start_h, sch_start_m = map(int, school.start_time.split(':'))
            start_dt = datetime.combine(now.date(), time(sch_start_h, sch_start_m))
            
            # Use school duration (default 60 mins if not set)
            dur = school.duration if school.duration else 60
            end_dt = start_dt + timedelta(minutes=dur)
            
            # 1. Check if Too Early
            if now < start_dt:
                return jsonify({"status": "error", "message": f"Too Early! Starts at {school.start_time}"}), 403
                
            # 2. Check if Time is Over
            if now > end_dt:
                # Trigger Auto Absent for the whole school
                trigger_auto_absent(school.id)
                return jsonify({"status": "error", "message": "Time Over - Marked Absent"}), 403

        except Exception as e:
            print(f"Time parsing error: {e}")
            # If time fails, allow attendance but log error (Fallback)
        
        # 3. If Time is OK, Save Log
        db.session.add(AttendanceLog(
            student_name=student.name, 
            student_uid=uid, 
            status=input_status, 
            place=data.get('place'), 
            school_id=school.id
        ))
        db.session.commit()
        
        # Update Sync Timestamp
        global DATA_LAST_UPDATED
        DATA_LAST_UPDATED = datetime.now().timestamp()
        
        return jsonify({"status": "success", "message": "Attendance Marked"}), 200
        
    except Exception as e:
        print(e)
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/stats/school', methods=['GET'])
@login_required
def get_school_stats_api():
    if current_user.role != 'super_admin': return jsonify({'error': 'Unauthorized'}), 403
    schools = School.query.all()
    stats = []
    today_str = str(date.today())
    for sch in schools:
        total = Student.query.filter_by(school_id=sch.id).count()
        present = db.session.query(AttendanceLog.student_uid).filter(
            AttendanceLog.school_id == sch.id, AttendanceLog.status == 'PRESENT',
            func.date(AttendanceLog.timestamp) == today_str
        ).distinct().count()
        percentage = round((present / total) * 100, 1) if total > 0 else 0
        stats.append({'name': sch.name, 'percentage': percentage})
    return jsonify(stats)

@app.route('/api/stats/class/<int:school_id>', methods=['GET'])
@login_required
def get_class_stats_api(school_id):
    if current_user.role != 'principal' and current_user.school_id != school_id: return jsonify({'error': 'Unauthorized'}), 403
    classes = db.session.query(Student.classroom).filter_by(school_id=school_id).distinct().all()
    stats = []
    today_str = str(date.today())
    for cls_tuple in classes:
        cls_name = cls_tuple[0]
        total = Student.query.filter_by(school_id=school_id, classroom=cls_name).count()
        present = db.session.query(AttendanceLog.student_uid).join(Student, AttendanceLog.student_uid == Student.rfid_uid).filter(
            AttendanceLog.school_id == school_id, Student.classroom == cls_name,
            AttendanceLog.status == 'PRESENT', func.date(AttendanceLog.timestamp) == today_str
        ).distinct().count()
        percentage = round((present / total) * 100, 1) if total > 0 else 0
        stats.append({'name': f"Class {cls_name}", 'percentage': percentage})
    return jsonify(stats)

def init_db():
    if not User.query.filter_by(username='admin').first():
        hashed_pw = bcrypt.generate_password_hash('admin123').decode('utf-8')
        db.session.add(User(username='admin', password=hashed_pw, role='super_admin'))
        db.session.commit()

# --- SERVER START ---
if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        init_db()
    
    print("------------------------------------------------")
    print(" SERVER STARTING (SQLite Mode)")
    print("------------------------------------------------")
    
    app.run(host='0.0.0.0', port=5000, debug=True)